import sys
import json
import os
from datetime import datetime, timedelta
from PyQt5.QtWidgets import (QApplication, QWidget, QPushButton, QHBoxLayout,
                             QVBoxLayout, QLabel, QLineEdit, QListWidget,
                             QDialog, QDialogButtonBox, QMessageBox, QFileDialog,
                             QSlider, QMenu, QAction, QWidgetAction,QSystemTrayIcon)
from PyQt5.QtCore import Qt, QTimer, QPoint
from PyQt5.QtGui import QFont, QColor, QPalette
from PyQt5.QtGui import QIcon, QPixmap, QPainter, QBrush
import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime
import shutil
from PyQt5.QtWidgets import QSystemTrayIcon, QMenu as QSystemTrayMenu

def get_icon_path():
    """获取图标路径（兼容开发环境和打包后的exe）"""
    if getattr(sys, 'frozen', False):
        base = sys._MEIPASS
    else:
        base = get_base_path()
    return os.path.join(base, "shiji.ico")

def get_base_path():
    """获取程序所在目录（兼容开发环境和打包后的exe）"""
    if getattr(sys, 'frozen', False):
        # 如果是打包后的 exe，返回 exe 所在的目录
        return os.path.dirname(sys.executable)
    else:
        # 如果是开发环境（.py），返回 .py 文件所在的目录
        return os.path.dirname(os.path.abspath(__file__))
    
def get_config_dir():
    """获取用户配置目录（跨平台），并确保目录存在"""
    if sys.platform == 'win32':
        base_dir = os.getenv('APPDATA')  # C:\Users\用户名\AppData\Roaming
    elif sys.platform == 'darwin':
        base_dir = os.path.expanduser('~/Library/Application Support')
    else:  # Linux and others
        base_dir = os.path.expanduser('~/.config')
    config_dir = os.path.join(base_dir, 'ShiJi')
    os.makedirs(config_dir, exist_ok=True)
    return config_dir



# ---------- 自定义任务选择对话框 ----------
class TaskSelectDialog(QDialog):
    def __init__(self, task_list, current_task, parent=None):
        super().__init__(parent)
        self.parent_window = parent  # 保存父窗口引用，方便操作数据
        self.setWindowTitle("选择或新建任务")
        self.setModal(True)
        self.resize(420, 350)

        font = QFont("Microsoft YaHei", 12)
        self.setFont(font)

        layout = QVBoxLayout()

        # 提示标签
        label = QLabel("点击任务名选择，或在下方输入新名称：")
        layout.addWidget(label)

        # 任务列表
        self.list_widget = QListWidget()
        self.list_widget.addItems(task_list)
        if current_task and current_task in task_list:
            self.list_widget.setCurrentRow(task_list.index(current_task))
        self.list_widget.itemDoubleClicked.connect(self.accept_selection)
        layout.addWidget(self.list_widget)

        # 操作行：输入框 + 新建按钮 + 删除按钮
        h_layout = QHBoxLayout()
        self.new_task_edit = QLineEdit()
        self.new_task_edit.setPlaceholderText("输入新任务名...")
        h_layout.addWidget(self.new_task_edit)

        # 新建按钮 - 绿色正向引导
        add_btn = QPushButton("新建")
        add_btn.clicked.connect(self.add_new_task)
        add_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 6px 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #43A047;
            }
            QPushButton:pressed {
                background-color: #388E3C;
            }
        """)
        h_layout.addWidget(add_btn)

        self.delete_btn = QPushButton("删除")
        self.delete_btn.clicked.connect(self.delete_selected_task)
        self.delete_btn.setStyleSheet("""
            QPushButton {
                background-color: #B33A3A;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 6px 14px;
            }
            QPushButton:hover {
                background-color: #9E3030;
            }
            QPushButton:pressed {
                background-color: #802828;
            }
        """)
        h_layout.addWidget(self.delete_btn)

        layout.addLayout(h_layout)

        # 标准按钮（确定/取消）
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        # 获取 OK 按钮并设置为引导蓝
        ok_btn = button_box.button(QDialogButtonBox.Ok)
        ok_btn.setText("确认")
        ok_btn.setStyleSheet("""
            QPushButton {
                background-color: #0078D7;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 30px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #106EBE;
            }
            QPushButton:pressed {
                background-color: #005A9E;
            }
        """)
        # 将 Cancel 按钮改为中文“取消”，并应用基础样式
        cancel_btn = button_box.button(QDialogButtonBox.Cancel)
        cancel_btn.setText("取消")
        cancel_btn.setStyleSheet("""
            QPushButton {
                border: none;
                border-radius: 4px;
                padding: 6px 20px;
                background-color: #E0E0E0;
                color: #333333;
            }
            QPushButton:hover {
                background-color: #D0D0D0;
            }
            QPushButton:pressed {
                background-color: #C0C0C0;
            }
        """)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
        self.setLayout(layout)
        self.selected_task = current_task

    def add_new_task(self):
        new_name = self.new_task_edit.text().strip()
        if new_name:
            items = [self.list_widget.item(i).text() for i in range(self.list_widget.count())]
            if new_name not in items:
                self.list_widget.addItem(new_name)
                self.list_widget.setCurrentRow(self.list_widget.count() - 1)
            else:
                self.list_widget.setCurrentRow(items.index(new_name))
            self.new_task_edit.clear()

    def delete_selected_task(self):
        """删除当前选中的任务，窗口保持打开，可连续删除"""
        current_item = self.list_widget.currentItem()
        if not current_item:
            QMessageBox.information(self, "提示", "请先选中一个任务再删除")
            return

        task_name = current_item.text()

        # 确认删除
        reply = QMessageBox.question(
            self,
            "确认删除",
            f"确定要删除任务「{task_name}」吗？\n（已用该任务名记录的工时不会受影响）",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return

        row = self.list_widget.currentRow()
        self.list_widget.takeItem(row)

        # 从主窗口任务列表中移除
        if task_name in self.parent_window.task_names:
            self.parent_window.task_names.remove(task_name)
            self.parent_window.save_data()  # 即时保存

        # 如果当前正在使用的任务被删了，更新主窗口按钮
        if self.parent_window.current_task == task_name:
            if self.list_widget.count() > 0:
                new_current = self.list_widget.item(0).text()
                self.parent_window.current_task = new_current
                self.parent_window.task_btn.setText(
                    new_current[:10] + ".." if len(new_current) > 10 else new_current
                )
            else:
                self.parent_window.current_task = ""
                self.parent_window.task_btn.setText("选择任务")

        # 如果列表还有内容，自动选中第一项
        if self.list_widget.count() > 0:
            self.list_widget.setCurrentRow(0)

    def accept_selection(self, item):
        self.selected_task = item.text()
        self.accept()

    def accept(self):
        current_item = self.list_widget.currentItem()
        if current_item:
            self.selected_task = current_item.text()
        else:
            text = self.new_task_edit.text().strip()
            if text:
                self.selected_task = text
            else:
                QMessageBox.warning(self, "提示", "请选择一个任务或输入新名称")
                return
        super().accept()

# ---------- 主窗口 ----------
class WorkTimer(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowFlags(Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint)
        self.setAttribute(Qt.WA_TranslucentBackground)  # 恢复透明背景
        self.setWindowOpacity(0.9)

        # 数据
        self.task_names = []
        self.current_task = ""
        self.start_time = None
        self.elapsed_seconds = 0
        self.is_running = False
        self.records = []

        self.init_ui()
        self.load_data()
        # 托盘初始化在 init_ui 末尾已经调用

    def init_ui(self):
        # 全局字体调大
        font = QFont("Microsoft YaHei", 14)
        self.setFont(font)

        layout = QHBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(20, 10, 20, 10)

        # 1. 任务名称按钮（点击弹出对话框）
        self.task_btn = QPushButton("选择任务")
        self.task_btn.setFixedSize(120, 60)
        self.task_btn.setStyleSheet("""
            QPushButton {
                border-radius: 30px;
                background-color: #607D8B;
                color: white;
                font-size: 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #546E7A;
            }
        """)
        self.task_btn.clicked.connect(self.show_task_dialog)
        layout.addWidget(self.task_btn)

        # 2. 开始/暂停
        self.start_btn = QPushButton("▶")
        self.start_btn.setFixedSize(60, 60)
        self.start_btn.setStyleSheet("""
            QPushButton {
                border-radius: 30px;
                background-color: #4CAF50;
                color: white;
                font-size: 24px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        self.start_btn.clicked.connect(self.toggle_timer)
        layout.addWidget(self.start_btn)

        # 3. 时间显示
        self.time_label = QLabel("00:00:00")
        self.time_label.setFixedSize(220, 60)
        self.time_label.setAlignment(Qt.AlignCenter)
        self.time_label.setFont(QFont("Consolas", 22, QFont.Bold))
        self.time_label.setStyleSheet("""
            QLabel {
                background-color: rgba(0, 0, 0, 180);
                color: #00FF00;
                border-radius: 30px;
                padding: 0 15px;
            }
        """)
        layout.addWidget(self.time_label)

        # 4. 结束任务
        self.end_btn = QPushButton("⏹")
        self.end_btn.setFixedSize(60, 60)
        self.end_btn.setStyleSheet("""
            QPushButton {
                border-radius: 30px;
                background-color: #f44336;
                color: white;
                font-size: 24px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #da190b;
            }
        """)
        self.end_btn.clicked.connect(self.end_task)
        layout.addWidget(self.end_btn)

        # 5. 保存记录
        self.save_btn = QPushButton("💾")
        self.save_btn.setFixedSize(60, 60)
        self.save_btn.setStyleSheet("""
            QPushButton {
                border-radius: 30px;
                background-color: #2196F3;
                color: white;
                font-size: 24px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #0b7dda;
            }
        """)
        self.save_btn.clicked.connect(self.save_records)
        layout.addWidget(self.save_btn)

        # 主容器
        self.main_widget = QWidget()
        self.main_widget.setLayout(layout)
        self.main_widget.setStyleSheet("""
            QWidget {
                background-color: rgba(50, 50, 50, 200);
                border-radius: 30px;
            }
        """)
        main_layout = QVBoxLayout()
        main_layout.addWidget(self.main_widget)
        self.setLayout(main_layout)

        # 计时器
        self.timer = QTimer()
        self.timer.timeout.connect(self.update_time)

        # 窗口拖动
        self.old_pos = None
        self.main_widget.mousePressEvent = self.mouse_press
        self.main_widget.mouseMoveEvent = self.mouse_move

        # 右键菜单
        self.setContextMenuPolicy(Qt.CustomContextMenu)
        self.customContextMenuRequested.connect(self.show_context_menu)
        # 系统托盘
        self.tray_icon = None
        self.tray_menu = None
        self.init_tray()
        # 自定义关闭按钮
        self.close_btn = QPushButton("✕", self)
        self.close_btn.setFixedSize(30, 30)
        self.close_btn.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                color: #aaaaaa;
                font-size: 20px;
                border: none;
                border-radius: 15px;
            }
            QPushButton:hover {
                background-color: #ff4444;
                color: white;
            }
        """)
        self.close_btn.clicked.connect(self.close)
        # 位置在 resizeEvent 中设定

    def resizeEvent(self, event):
        self.close_btn.move(self.width() - 40, 10)
        super().resizeEvent(event)

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_F4 and event.modifiers() == Qt.AltModifier:
            self.close()
            event.accept()
        else:
            super().keyPressEvent(event)

    def init_tray(self):
        """初始化系统托盘"""
        icon_path = get_icon_path()
        if not os.path.exists(icon_path):
            icon = QIcon()
        else:
            icon = QIcon(icon_path)

        self.tray_icon = QSystemTrayIcon(icon, self)
        self.tray_icon.setToolTip("时迹 - 点击显示/隐藏")
        self.tray_icon.activated.connect(self.on_tray_activated)

        self.tray_menu = QSystemTrayMenu()
        show_action = QAction("显示主窗口", self)
        show_action.triggered.connect(self.show_window)
        self.tray_menu.addAction(show_action)

        hide_action = QAction("隐藏主窗口", self)
        hide_action.triggered.connect(self.hide_window)
        self.tray_menu.addAction(hide_action)

        self.tray_menu.addSeparator()

        self.autostart_action = QAction("开机自启", self)
        self.autostart_action.setCheckable(True)
        self.autostart_action.setChecked(self.is_autostart_enabled())
        self.autostart_action.triggered.connect(self.toggle_autostart)
        self.tray_menu.addAction(self.autostart_action)

        self.tray_menu.addSeparator()

        quit_action = QAction("退出", self)
        quit_action.triggered.connect(self.quit_app)
        self.tray_menu.addAction(quit_action)

        self.tray_icon.setContextMenu(self.tray_menu)
        self.tray_icon.show()

    def on_tray_activated(self, reason):
        if reason == QSystemTrayIcon.Trigger:
            if self.isVisible():
                self.hide()
            else:
                self.show()
                self.raise_()
                self.activateWindow()

    def show_window(self):
        self.show()
        self.raise_()
        self.activateWindow()

    def hide_window(self):
        self.hide()

    def quit_app(self):
        self.tray_icon.hide()
        QApplication.quit()

    def closeEvent(self, event):
        if self.tray_icon and self.tray_icon.isVisible():
            self.hide()
            event.ignore()
        else:
            event.accept()

    def is_autostart_enabled(self):
        if sys.platform != 'win32':
            return False
        try:
            import winreg
            key = winreg.OpenKey(
                winreg.HKEY_CURRENT_USER,
                r"Software\Microsoft\Windows\CurrentVersion\Run",
                0,
                winreg.KEY_READ
            )
            try:
                value, _ = winreg.QueryValueEx(key, "ShiJi")
                winreg.CloseKey(key)
                current_path = sys.executable if getattr(sys, 'frozen', False) else __file__
                return value == f'"{current_path}"'
            except FileNotFoundError:
                winreg.CloseKey(key)
                return False
        except Exception:
            return False

    def toggle_autostart(self):
        if sys.platform != 'win32':
            QMessageBox.information(self, "提示", "当前系统暂不支持开机自启设置")
            return
        try:
            import winreg
            key = winreg.OpenKey(
                winreg.HKEY_CURRENT_USER,
                r"Software\Microsoft\Windows\CurrentVersion\Run",
                0,
                winreg.KEY_SET_VALUE | winreg.KEY_READ
            )
            current_path = sys.executable if getattr(sys, 'frozen', False) else __file__
            if self.autostart_action.isChecked():
                winreg.SetValueEx(key, "ShiJi", 0, winreg.REG_SZ, f'"{current_path}"')
            else:
                try:
                    winreg.DeleteValue(key, "ShiJi")
                except FileNotFoundError:
                    pass
            winreg.CloseKey(key)
        except Exception as e:
            QMessageBox.warning(self, "错误", f"修改开机自启失败：{str(e)}")
            self.autostart_action.setChecked(self.is_autostart_enabled())



    # ---------- 任务选择对话框 ----------
    def show_task_dialog(self):
        dialog = TaskSelectDialog(self.task_names, self.current_task, self)
        if dialog.exec_() == QDialog.Accepted:
            new_task = dialog.selected_task
            if new_task and new_task != self.current_task:
                self.current_task = new_task
                self.task_btn.setText(new_task[:10] + ".." if len(new_task) > 10 else new_task)
                # 如果新任务不在列表中，添加到列表并保存
                if new_task not in self.task_names:
                    self.task_names.append(new_task)
                    self.save_data()

    # ---------- 其他功能（计时、结束、保存等） ----------
    def toggle_timer(self):
        if not self.current_task:
            QMessageBox.warning(self, "提示", "请先选择或新建一个任务！")
            return
        if self.is_running:
            self.timer.stop()
            self.is_running = False
            self.start_btn.setText("▶")
            self.start_btn.setStyleSheet("""
                QPushButton {
                    border-radius: 30px;
                    background-color: #4CAF50;
                    color: white;
                    font-size: 24px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #45a049;
                }
            """)
        else:
            if self.start_time is None:
                self.start_time = datetime.now()
            else:
                self.start_time = datetime.now() - timedelta(seconds=self.elapsed_seconds)
            self.timer.start(1000)
            self.is_running = True
            self.start_btn.setText("⏸")
            self.start_btn.setStyleSheet("""
                QPushButton {
                    border-radius: 30px;
                    background-color: #FF9800;
                    color: white;
                    font-size: 24px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #e68a00;
                }
            """)

    def update_time(self):
        if self.start_time:
            delta = datetime.now() - self.start_time
            self.elapsed_seconds = int(delta.total_seconds())
            hours = self.elapsed_seconds // 3600
            minutes = (self.elapsed_seconds % 3600) // 60
            seconds = self.elapsed_seconds % 60
            self.time_label.setText(f"{hours:02d}:{minutes:02d}:{seconds:02d}")

    def end_task(self):
        if self.is_running:
            self.timer.stop()
            self.is_running = False
            self.start_btn.setText("▶")
            self.start_btn.setStyleSheet("""
                QPushButton {
                    border-radius: 30px;
                    background-color: #4CAF50;
                    color: white;
                    font-size: 24px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #45a049;
                }
            """)
        if self.elapsed_seconds > 0 and self.current_task:
            end_time = datetime.now()
            self.records.append((
                self.current_task,
                self.start_time.strftime("%Y-%m-%d %H:%M:%S"),
                end_time.strftime("%Y-%m-%d %H:%M:%S"),
                self.elapsed_seconds
            ))
            # 先显示“已记录”的视觉反馈（无声）
            self.time_label.setText("✔ 已记录")
            # 重置计时数据
            self.start_time = None
            self.elapsed_seconds = 0
            # 1.5秒后恢复时间显示
            QTimer.singleShot(1500, lambda: self.time_label.setText("00:00:00"))

    def save_records(self):
        if not self.records:
            self.time_label.setText("⚠ 无记录")
            QTimer.singleShot(1500, lambda: self.time_label.setText("00:00:00"))
            return

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_name = f"工作时间记录_{timestamp}.xlsx"
        desktop = str(Path.home() / "Desktop")
        
        # 处理重名
        file_path = os.path.join(desktop, base_name)
        counter = 1
        while os.path.exists(file_path):
            name_without_ext = f"工作时间记录_{timestamp}({counter})"
            file_path = os.path.join(desktop, f"{name_without_ext}.xlsx")
            counter += 1

        # 如果你希望每次弹出对话框让用户确认位置，取消下面三行的注释
        file_path, _ = QFileDialog.getSaveFileName(self, "保存记录", file_path, "Excel文件 (*.xlsx)")
        if not file_path:
             return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "工作时间记录"
            ws.append(["任务名称", "开始时间", "结束时间", "总时长"])
            for task, start, end, seconds in self.records:
                hours = seconds // 3600
                minutes = (seconds % 3600) // 60
                secs = seconds % 60
                time_str = f"{hours:02d}:{minutes:02d}:{secs:02d}"
                ws.append([task, start, end, time_str])
            column_widths = [10, 22, 22, 10]
            for i, width in enumerate(column_widths, start=1):
                col_letter = get_column_letter(i)
                ws.column_dimensions[col_letter].width = width
            wb.save(file_path)
            self.time_label.setText("✔ 已保存")
            QTimer.singleShot(1500, lambda: self.time_label.setText("00:00:00"))
        except Exception as e:
            self.time_label.setText("✘ 保存失败")
            QTimer.singleShot(1500, lambda: self.time_label.setText("00:00:00"))

    # ---------- 数据持久化 ----------
    def load_data(self):
        config_dir = get_config_dir()
        new_file = os.path.join(config_dir, "tasks.json")
        old_file = os.path.join(get_base_path(), "tasks.json")

        # 如果新目录有文件，直接用；否则尝试从旧目录迁移
        if os.path.exists(new_file):
            file_path = new_file
        elif os.path.exists(old_file):
            # 自动迁移旧数据
            import shutil
            shutil.copy2(old_file, new_file)
            # 迁移后保留旧文件作为备份（不删除）
            file_path = new_file
        else:
            # 都没有，直接返回（首次启动）
            return

        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)
            self.task_names = data.get("tasks", [])
            if self.task_names:
                self.current_task = self.task_names[0]
                self.task_btn.setText(self.current_task[:10] + ".." if len(self.current_task) > 10 else self.current_task)

    def save_data(self):
        config_dir = get_config_dir()
        file_path = os.path.join(config_dir, "tasks.json")
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump({"tasks": self.task_names}, f, ensure_ascii=False, indent=2)
        
    # ---------- 窗口拖动 ----------
    def mouse_press(self, event):
        if event.button() == Qt.LeftButton:
            self.old_pos = event.globalPos()

    def mouse_move(self, event):
        if self.old_pos is not None:
            delta = event.globalPos() - self.old_pos
            self.move(self.x() + delta.x(), self.y() + delta.y())
            self.old_pos = event.globalPos()

    def mouseReleaseEvent(self, event):
        self.old_pos = None

    # ---------- 右键菜单 ----------
    def show_context_menu(self, pos):
        menu = QMenu(self)

        opacity_action = QWidgetAction(menu)
        slider = QSlider(Qt.Horizontal)
        slider.setRange(10, 100)
        slider.setValue(int(self.windowOpacity() * 100))
        slider.setFixedWidth(150)
        slider.valueChanged.connect(lambda v: self.setWindowOpacity(v / 100))
        opacity_action.setDefaultWidget(slider)
        menu.addAction(opacity_action)

        pin_action = QAction("始终置顶" if self.windowFlags() & Qt.WindowStaysOnTopHint else "取消置顶")
        pin_action.triggered.connect(self.toggle_pin)
        menu.addAction(pin_action)

        menu.exec_(self.mapToGlobal(pos))

    def toggle_pin(self):
        if self.windowFlags() & Qt.WindowStaysOnTopHint:
            self.setWindowFlags(self.windowFlags() & ~Qt.WindowStaysOnTopHint)
        else:
            self.setWindowFlags(self.windowFlags() | Qt.WindowStaysOnTopHint)
        self.show()

if __name__ == "__main__":
    import sys
    from PyQt5.QtCore import QSharedMemory

    app = QApplication(sys.argv)

    # 设置全局应用图标（任务栏 + 窗口）
    app.setWindowIcon(QIcon(get_icon_path()))

    # 单实例检测
    shared_memory = QSharedMemory("Task_Timer_Single_Instance_54_iKun")
    if shared_memory.attach():
        print("程序已在运行中，不能重复打开")
        sys.exit(0)
    else:
        shared_memory.create(1)

    window = WorkTimer()
    window.show()
    sys.exit(app.exec_())