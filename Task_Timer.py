import sys
import json
import os
from datetime import datetime, timedelta
from PyQt5.QtWidgets import (QApplication, QWidget, QPushButton, QHBoxLayout,
                             QVBoxLayout, QLabel, QLineEdit, QListWidget,
                             QDialog, QDialogButtonBox, QMessageBox, QFileDialog,
                             QSlider, QMenu, QAction, QWidgetAction)
from PyQt5.QtCore import Qt, QTimer, QPoint
from PyQt5.QtGui import QFont, QColor, QPalette
import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime

def get_base_path():
    """获取程序所在目录（兼容开发环境和打包后的exe）"""
    if getattr(sys, 'frozen', False):
        # 如果是打包后的 exe，返回 exe 所在的目录
        return os.path.dirname(sys.executable)
    else:
        # 如果是开发环境（.py），返回 .py 文件所在的目录
        return os.path.dirname(os.path.abspath(__file__))

# ---------- 自定义任务选择对话框 ----------
class TaskSelectDialog(QDialog):
    def __init__(self, task_list, current_task, parent=None):
        super().__init__(parent)
        self.setWindowTitle("选择或新建任务")
        self.setModal(True)
        self.resize(400, 300)
        
        # 字体调大
        font = QFont("Microsoft YaHei", 12)
        self.setFont(font)
        
        layout = QVBoxLayout()
        
        # 提示标签
        label = QLabel("点击任务名选择，或在下方输入新名称：")
        layout.addWidget(label)
        
        # 任务列表
        self.list_widget = QListWidget()
        self.list_widget.addItems(task_list)
        if current_task and current_task in task_list:
            self.list_widget.setCurrentRow(task_list.index(current_task))
        self.list_widget.itemDoubleClicked.connect(self.accept_selection)  # 双击选中
        layout.addWidget(self.list_widget)
        
        # 新建任务输入框和按钮
        h_layout = QHBoxLayout()
        self.new_task_edit = QLineEdit()
        self.new_task_edit.setPlaceholderText("输入新任务名...")
        h_layout.addWidget(self.new_task_edit)
        add_btn = QPushButton("新建")
        add_btn.clicked.connect(self.add_new_task)
        h_layout.addWidget(add_btn)
        layout.addLayout(h_layout)
        
        # 标准按钮（确定/取消）
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
        
        self.setLayout(layout)
        self.selected_task = current_task  # 默认当前任务
        
    def add_new_task(self):
        new_name = self.new_task_edit.text().strip()
        if new_name:
            # 如果列表中已有，直接选中；否则添加
            items = [self.list_widget.item(i).text() for i in range(self.list_widget.count())]
            if new_name not in items:
                self.list_widget.addItem(new_name)
                self.list_widget.setCurrentRow(self.list_widget.count() - 1)
            else:
                # 已有则选中
                self.list_widget.setCurrentRow(items.index(new_name))
            self.new_task_edit.clear()
    
    def accept_selection(self, item):
        self.selected_task = item.text()
        self.accept()
    
    def accept(self):
        # 确定时取当前选中项
        current_item = self.list_widget.currentItem()
        if current_item:
            self.selected_task = current_item.text()
        else:
            # 如果列表为空且输入框有内容，使用输入框
            text = self.new_task_edit.text().strip()
            if text:
                self.selected_task = text
            else:
                QMessageBox.warning(self, "提示", "请选择一个任务或输入新名称")
                return
        super().accept()

# ---------- 主窗口 ----------
class WorkTimer(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowFlags(Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint)
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.setWindowOpacity(0.9)

        # 数据
        self.task_names = []
        self.current_task = ""
        self.start_time = None
        self.elapsed_seconds = 0
        self.is_running = False
        self.records = []

        self.init_ui()
        self.load_data()

    def init_ui(self):
        # 全局字体调大
        font = QFont("Microsoft YaHei", 14)
        self.setFont(font)

        layout = QHBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(20, 10, 20, 10)

        # 1. 任务名称按钮（点击弹出对话框）
        self.task_btn = QPushButton("选择任务")
        self.task_btn.setFixedSize(120, 60)
        self.task_btn.setStyleSheet("""
            QPushButton {
                border-radius: 30px;
                background-color: #607D8B;
                color: white;
                font-size: 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #546E7A;
            }
        """)
        self.task_btn.clicked.connect(self.show_task_dialog)
        layout.addWidget(self.task_btn)

        # 2. 开始/暂停
        self.start_btn = QPushButton("▶")
        self.start_btn.setFixedSize(60, 60)
        self.start_btn.setStyleSheet("""
            QPushButton {
                border-radius: 30px;
                background-color: #4CAF50;
                color: white;
                font-size: 24px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        self.start_btn.clicked.connect(self.toggle_timer)
        layout.addWidget(self.start_btn)

        # 3. 时间显示
        self.time_label = QLabel("00:00:00")
        self.time_label.setFixedSize(220, 60)
        self.time_label.setAlignment(Qt.AlignCenter)
        self.time_label.setFont(QFont("Consolas", 22, QFont.Bold))
        self.time_label.setStyleSheet("""
            QLabel {
                background-color: rgba(0, 0, 0, 180);
                color: #00FF00;
                border-radius: 30px;
                padding: 0 15px;
            }
        """)
        layout.addWidget(self.time_label)

        # 4. 结束任务
        self.end_btn = QPushButton("⏹")
        self.end_btn.setFixedSize(60, 60)
        self.end_btn.setStyleSheet("""
            QPushButton {
                border-radius: 30px;
                background-color: #f44336;
                color: white;
                font-size: 24px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #da190b;
            }
        """)
        self.end_btn.clicked.connect(self.end_task)
        layout.addWidget(self.end_btn)

        # 5. 保存记录
        self.save_btn = QPushButton("💾")
        self.save_btn.setFixedSize(60, 60)
        self.save_btn.setStyleSheet("""
            QPushButton {
                border-radius: 30px;
                background-color: #2196F3;
                color: white;
                font-size: 24px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #0b7dda;
            }
        """)
        self.save_btn.clicked.connect(self.save_records)
        layout.addWidget(self.save_btn)

        # 主容器
        self.main_widget = QWidget()
        self.main_widget.setLayout(layout)
        self.main_widget.setStyleSheet("""
            QWidget {
                background-color: rgba(50, 50, 50, 200);
                border-radius: 30px;
            }
        """)
        main_layout = QVBoxLayout()
        main_layout.addWidget(self.main_widget)
        self.setLayout(main_layout)

        # 计时器
        self.timer = QTimer()
        self.timer.timeout.connect(self.update_time)

        # 窗口拖动
        self.old_pos = None
        self.main_widget.mousePressEvent = self.mouse_press
        self.main_widget.mouseMoveEvent = self.mouse_move

        # 右键菜单
        self.setContextMenuPolicy(Qt.CustomContextMenu)
        self.customContextMenuRequested.connect(self.show_context_menu)

    # ---------- 任务选择对话框 ----------
    def show_task_dialog(self):
        dialog = TaskSelectDialog(self.task_names, self.current_task, self)
        if dialog.exec_() == QDialog.Accepted:
            new_task = dialog.selected_task
            if new_task and new_task != self.current_task:
                self.current_task = new_task
                self.task_btn.setText(new_task[:10] + ".." if len(new_task) > 10 else new_task)
                # 如果新任务不在列表中，添加到列表并保存
                if new_task not in self.task_names:
                    self.task_names.append(new_task)
                    self.save_data()

    # ---------- 其他功能（计时、结束、保存等） ----------
    def toggle_timer(self):
        if not self.current_task:
            QMessageBox.warning(self, "提示", "请先选择或新建一个任务！")
            return
        if self.is_running:
            self.timer.stop()
            self.is_running = False
            self.start_btn.setText("▶")
            self.start_btn.setStyleSheet("""
                QPushButton {
                    border-radius: 30px;
                    background-color: #4CAF50;
                    color: white;
                    font-size: 24px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #45a049;
                }
            """)
        else:
            if self.start_time is None:
                self.start_time = datetime.now()
            else:
                self.start_time = datetime.now() - timedelta(seconds=self.elapsed_seconds)
            self.timer.start(1000)
            self.is_running = True
            self.start_btn.setText("⏸")
            self.start_btn.setStyleSheet("""
                QPushButton {
                    border-radius: 30px;
                    background-color: #FF9800;
                    color: white;
                    font-size: 24px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #e68a00;
                }
            """)

    def update_time(self):
        if self.start_time:
            delta = datetime.now() - self.start_time
            self.elapsed_seconds = int(delta.total_seconds())
            hours = self.elapsed_seconds // 3600
            minutes = (self.elapsed_seconds % 3600) // 60
            seconds = self.elapsed_seconds % 60
            self.time_label.setText(f"{hours:02d}:{minutes:02d}:{seconds:02d}")

    def end_task(self):
        if self.is_running:
            self.timer.stop()
            self.is_running = False
            self.start_btn.setText("▶")
            self.start_btn.setStyleSheet("""
                QPushButton {
                    border-radius: 30px;
                    background-color: #4CAF50;
                    color: white;
                    font-size: 24px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #45a049;
                }
            """)
        if self.elapsed_seconds > 0 and self.current_task:
            end_time = datetime.now()
            self.records.append((
                self.current_task,
                self.start_time.strftime("%Y-%m-%d %H:%M:%S"),
                end_time.strftime("%Y-%m-%d %H:%M:%S"),
                self.elapsed_seconds
            ))
            # 先显示“已记录”的视觉反馈（无声）
            self.time_label.setText("✔ 已记录")
            # 重置计时数据
            self.start_time = None
            self.elapsed_seconds = 0
            # 1.5秒后恢复时间显示
            QTimer.singleShot(1500, lambda: self.time_label.setText("00:00:00"))

    def save_records(self):
        if not self.records:
            self.time_label.setText("⚠ 无记录")
            QTimer.singleShot(1500, lambda: self.time_label.setText("00:00:00"))
            return

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_name = f"工作时间记录_{timestamp}.xlsx"
        desktop = str(Path.home() / "Desktop")
        
        # 处理重名
        file_path = os.path.join(desktop, base_name)
        counter = 1
        while os.path.exists(file_path):
            name_without_ext = f"工作时间记录_{timestamp}({counter})"
            file_path = os.path.join(desktop, f"{name_without_ext}.xlsx")
            counter += 1

        # 如果你希望每次弹出对话框让用户确认位置，取消下面三行的注释
        file_path, _ = QFileDialog.getSaveFileName(self, "保存记录", file_path, "Excel文件 (*.xlsx)")
        if not file_path:
             return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "工作时间记录"
            ws.append(["任务名称", "开始时间", "结束时间", "总时长"])
            for task, start, end, seconds in self.records:
                hours = seconds // 3600
                minutes = (seconds % 3600) // 60
                secs = seconds % 60
                time_str = f"{hours:02d}:{minutes:02d}:{secs:02d}"
                ws.append([task, start, end, time_str])
            column_widths = [10, 22, 22, 10]
            for i, width in enumerate(column_widths, start=1):
                col_letter = get_column_letter(i)
                ws.column_dimensions[col_letter].width = width
            wb.save(file_path)
            self.time_label.setText("✔ 已保存")
            QTimer.singleShot(1500, lambda: self.time_label.setText("00:00:00"))
        except Exception as e:
            self.time_label.setText("✘ 保存失败")
            QTimer.singleShot(1500, lambda: self.time_label.setText("00:00:00"))

    # ---------- 数据持久化 ----------
    def load_data(self):
        base_dir = get_base_path() 
        file_path = os.path.join(base_dir, "tasks.json")
        if os.path.exists(file_path):
            with open(file_path, "r", encoding="utf-8") as f:
                data = json.load(f)
                self.task_names = data.get("tasks", [])
                if self.task_names:
                    self.current_task = self.task_names[0]
                    self.task_btn.setText(self.current_task[:10] + ".." if len(self.current_task) > 10 else self.current_task)

    def save_data(self):
        base_dir = get_base_path()  # 替换原来的 script_dir
        file_path = os.path.join(base_dir, "tasks.json")
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump({"tasks": self.task_names}, f, ensure_ascii=False, indent=2)
    # ---------- 窗口拖动 ----------
    def mouse_press(self, event):
        if event.button() == Qt.LeftButton:
            self.old_pos = event.globalPos()

    def mouse_move(self, event):
        if self.old_pos is not None:
            delta = event.globalPos() - self.old_pos
            self.move(self.x() + delta.x(), self.y() + delta.y())
            self.old_pos = event.globalPos()

    def mouseReleaseEvent(self, event):
        self.old_pos = None

    # ---------- 右键菜单 ----------
    def show_context_menu(self, pos):
        menu = QMenu(self)

        opacity_action = QWidgetAction(menu)
        slider = QSlider(Qt.Horizontal)
        slider.setRange(10, 100)
        slider.setValue(int(self.windowOpacity() * 100))
        slider.setFixedWidth(150)
        slider.valueChanged.connect(lambda v: self.setWindowOpacity(v / 100))
        opacity_action.setDefaultWidget(slider)
        menu.addAction(opacity_action)

        pin_action = QAction("始终置顶" if self.windowFlags() & Qt.WindowStaysOnTopHint else "取消置顶")
        pin_action.triggered.connect(self.toggle_pin)
        menu.addAction(pin_action)

        menu.exec_(self.mapToGlobal(pos))

    def toggle_pin(self):
        if self.windowFlags() & Qt.WindowStaysOnTopHint:
            self.setWindowFlags(self.windowFlags() & ~Qt.WindowStaysOnTopHint)
        else:
            self.setWindowFlags(self.windowFlags() | Qt.WindowStaysOnTopHint)
        self.show()

if __name__ == "__main__":
    import sys
    from PyQt5.QtCore import QSharedMemory

    app = QApplication(sys.argv)

    # 使用 QSharedMemory 的 attach 方式检测
    shared_memory = QSharedMemory("Task_Timer_Single_Instance_54_iKun")
    if shared_memory.attach():
        # attach 成功说明已经有一个实例在运行了
        print("程序已在运行中，不能重复打开")
        sys.exit(0)
    else:
        # attach 失败，说明是第一个实例，创建共享内存
        shared_memory.create(1)

    window = WorkTimer()
    window.show()
    sys.exit(app.exec_())